from dataclasses import dataclass
from pathlib import Path
import logging

from openpyxl import load_workbook

from app.cases.models import CaseItem
from app.cases.repository import CaseRepository
from app.config import load_settings
from app.utils.text import build_search_text, normalize_column_name

logger = logging.getLogger(__name__)


COLUMN_ALIASES = {
    "project": {"проект", "project"},
    "title": {"название кейса", "кейс", "case", "case name", "title", "название"},
    "case_url": {"ссылка на кейс", "ссылка", "case url", "url", "link"},
    "niche": {"ниша", "category"},
    "niche_extra": {"ниша доп", "доп ниша", "ниша дополнительная", "niche extra"},
    "goal": {"цель", "goal"},
    "tool": {"инструмент", "tool", "канал", "channel"},
    "geo": {"гео", "география", "geo", "гео "},
    "audience": {"аудитория", "audience"},
    "storage_url": {"где лежат", "где лежит", "хранилище", "storage", "storage url"},
}


@dataclass
class ImportSummary:
    imported_count: int
    source_sheets: list[str]


def import_cases_from_excel(xlsx_path: Path, repository: CaseRepository) -> ImportSummary:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel-файл не найден: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    cases: list[CaseItem] = []
    seen_keys: set[tuple[str, str, str]] = set()

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue

        column_map = _build_column_map(headers)
        for row in rows:
            case = _row_to_case(sheet_name, row, column_map)
            if not _has_any_data(case):
                continue

            dedupe_key = (
                case.title.strip().lower(),
                case.case_url.strip().lower(),
                case.source_sheet.strip().lower(),
            )
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            cases.append(case)

    repository.replace_all(cases)
    logger.info("Импортировано кейсов: %s", len(cases))
    return ImportSummary(imported_count=len(cases), source_sheets=workbook.sheetnames)


def _build_column_map(columns: list[object]) -> dict[str, object]:
    column_map: dict[str, int] = {}

    for index, column in enumerate(columns):
        normalized_column = normalize_column_name(column)
        if not normalized_column:
            continue

        for field_name, aliases in COLUMN_ALIASES.items():
            normalized_aliases = {normalize_column_name(alias) for alias in aliases}
            if normalized_column in normalized_aliases and field_name not in column_map:
                column_map[field_name] = index
                break

    return column_map


def _row_to_case(sheet_name: str, row: tuple[object, ...], column_map: dict[str, int]) -> CaseItem:
    values = {field: _safe_cell(_row_value(row, index)) for field, index in column_map.items()}
    title = values.get("title", "")
    if not title:
        title = values.get("project", "") or values.get("case_url", "") or "Без названия"

    case = CaseItem(
        source_sheet=sheet_name,
        project=values.get("project", ""),
        title=title,
        case_url=values.get("case_url", ""),
        niche=values.get("niche", ""),
        niche_extra=values.get("niche_extra", ""),
        goal=values.get("goal", ""),
        tool=values.get("tool", ""),
        geo=values.get("geo", ""),
        audience=values.get("audience", ""),
        storage_url=values.get("storage_url", ""),
    )
    case.search_text = build_search_text(
        (
            case.project,
            case.title,
            case.case_url,
            case.niche,
            case.niche_extra,
            case.goal,
            case.tool,
            case.geo,
            case.audience,
            case.storage_url,
        )
    )
    return case


def _safe_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_value(row: tuple[object, ...], index: int) -> object:
    if index >= len(row):
        return None
    return row[index]


def _has_any_data(case: CaseItem) -> bool:
    return any(
        (
            case.project,
            case.title and case.title != "Без названия",
            case.case_url,
            case.niche,
            case.niche_extra,
            case.goal,
            case.tool,
            case.geo,
            case.audience,
            case.storage_url,
        )
    )


def main() -> None:
    settings = load_settings()
    repository = CaseRepository(settings.db_path)
    summary = import_cases_from_excel(settings.cases_xlsx_path, repository)
    print(f"Импортировано кейсов: {summary.imported_count}")
    print(f"Листы: {', '.join(summary.source_sheets)}")


if __name__ == "__main__":
    main()
